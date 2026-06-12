from tkinter import *
from tkinter import messagebox
from database import connect_db

from openpyxl import Workbook

from reportlab.platypus import (
    SimpleDocTemplate,
    Table,
    TableStyle,
    Paragraph,
    Spacer
)

from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet


def report_page(parent):

    frame = Frame(parent)

    # ===================================
    # EXPORT CUSTOMERS EXCEL
    # ===================================

    def export_customers_excel():

        try:

            conn = connect_db()
            cursor = conn.cursor()

            cursor.execute(
                """
                SELECT *
                FROM Customers
                """
            )

            rows = cursor.fetchall()

            wb = Workbook()

            ws = wb.active

            ws.title = "Customers"

            ws.append(
                [
                    "ID",
                    "Name",
                    "Email",
                    "Phone",
                    "City"
                ]
            )

            for row in rows:
                ws.append(row)

            wb.save(
                "Customer_Report.xlsx"
            )

            conn.close()

            messagebox.showinfo(
                "Success",
                "Customer_Report.xlsx Created"
            )

        except Exception as e:

            messagebox.showerror(
                "Error",
                str(e)
            )

    # ===================================
    # EXPORT PRODUCTS EXCEL
    # ===================================

    def export_products_excel():

        try:

            conn = connect_db()
            cursor = conn.cursor()

            cursor.execute(
                """
                SELECT *
                FROM Products
                """
            )

            rows = cursor.fetchall()

            wb = Workbook()

            ws = wb.active

            ws.title = "Products"

            ws.append(
                [
                    "ID",
                    "Product Name",
                    "Category",
                    "Price",
                    "Stock"
                ]
            )

            for row in rows:
                ws.append(row)

            wb.save(
                "Product_Report.xlsx"
            )

            conn.close()

            messagebox.showinfo(
                "Success",
                "Product_Report.xlsx Created"
            )

        except Exception as e:

            messagebox.showerror(
                "Error",
                str(e)
            )

    # ===================================
    # EXPORT SALES EXCEL
    # ===================================

    def export_sales_excel():

        try:

            conn = connect_db()
            cursor = conn.cursor()

            cursor.execute(
                """
                SELECT *
                FROM Orders
                """
            )

            rows = cursor.fetchall()

            wb = Workbook()

            ws = wb.active

            ws.title = "Sales"

            ws.append(
                [
                    "Order ID",
                    "Customer ID",
                    "Product ID",
                    "Quantity",
                    "Total",
                    "Date"
                ]
            )

            for row in rows:
                ws.append(row)

            wb.save(
                "Sales_Report.xlsx"
            )

            conn.close()

            messagebox.showinfo(
                "Success",
                "Sales_Report.xlsx Created"
            )

        except Exception as e:

            messagebox.showerror(
                "Error",
                str(e)
            )

    # ===================================
    # PDF REPORT
    # ===================================

    def export_sales_pdf():

        try:

            conn = connect_db()
            cursor = conn.cursor()

            cursor.execute(
                """
                SELECT
                order_id,
                customer_id,
                product_id,
                quantity,
                total_amount
                FROM Orders
                """
            )

            rows = cursor.fetchall()

            conn.close()

            pdf = SimpleDocTemplate(
                "Sales_Report.pdf"
            )

            styles = getSampleStyleSheet()

            elements = []

            title = Paragraph(
                "Sales Report",
                styles["Title"]
            )

            elements.append(title)

            elements.append(
                Spacer(1,12)
            )

            data = [
                [
                    "Order ID",
                    "Customer",
                    "Product",
                    "Qty",
                    "Total"
                ]
            ]

            for row in rows:
                data.append(list(row))

            table = Table(data)

            table.setStyle(
                TableStyle(
                    [
                        (
                            'BACKGROUND',
                            (0,0),
                            (-1,0),
                            colors.grey
                        ),
                        (
                            'TEXTCOLOR',
                            (0,0),
                            (-1,0),
                            colors.whitesmoke
                        ),
                        (
                            'GRID',
                            (0,0),
                            (-1,-1),
                            1,
                            colors.black
                        )
                    ]
                )
            )

            elements.append(table)

            pdf.build(elements)

            messagebox.showinfo(
                "Success",
                "Sales_Report.pdf Created"
            )

        except Exception as e:

            messagebox.showerror(
                "Error",
                str(e)
            )

    # ===================================
    # TITLE
    # ===================================

    Label(
        frame,
        text="Reports & Export Center",
        font=("Arial",20,"bold")
    ).pack(
        pady=20
    )

    # ===================================
    # EXCEL REPORTS
    # ===================================

    excel_frame = LabelFrame(
        frame,
        text="Excel Reports",
        padx=20,
        pady=20
    )

    excel_frame.pack(
        pady=10,
        fill=X,
        padx=20
    )

    Button(
        excel_frame,
        text="Export Customers Excel",
        width=30,
        command=export_customers_excel
    ).pack(
        pady=5
    )

    Button(
        excel_frame,
        text="Export Products Excel",
        width=30,
        command=export_products_excel
    ).pack(
        pady=5
    )

    Button(
        excel_frame,
        text="Export Sales Excel",
        width=30,
        command=export_sales_excel
    ).pack(
        pady=5
    )

    # ===================================
    # PDF REPORTS
    # ===================================

    pdf_frame = LabelFrame(
        frame,
        text="PDF Reports",
        padx=20,
        pady=20
    )

    pdf_frame.pack(
        pady=10,
        fill=X,
        padx=20
    )

    Button(
        pdf_frame,
        text="Export Sales PDF",
        width=30,
        command=export_sales_pdf
    ).pack(
        pady=5
    )

    # ===================================
    # PROJECT INFO
    # ===================================

    info = LabelFrame(
        frame,
        text="Project Features",
        padx=20,
        pady=20
    )

    info.pack(
        pady=10,
        fill=BOTH,
        expand=True,
        padx=20
    )

    Label(
        info,
        text="""
✓ Customer Management

✓ Product Management

✓ Order Management

✓ Revenue Dashboard

✓ Excel Export

✓ PDF Reports

✓ MySQL Connectivity

✓ Final Year Project
        """,
        justify=LEFT
    ).pack(anchor="w")

    return frame